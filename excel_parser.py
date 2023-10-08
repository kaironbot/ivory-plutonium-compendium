from openpyxl import load_workbook
import json
import os

DESC_SHEET = "Descrizione"
FEATURE_SHEET = "Feature Generiche"
SPELL_SHEET = "Incatesimi"
TALENTS_SHEET = "Talenti"
SPELL_PLACEHOLDER = "%SPELLS%"

def parse_resource_by_type(wb, type):
    if type == "Sottoclasse":
        return ["subclasses", parse_class(wb)]
    elif type == "Talento":
        return ["talent", parse_talent(wb)]
    return ["", {}]

def raw_features_to_feature(raw_feature, spells):
    ret = []
    for el in raw_feature:
        if el == SPELL_PLACEHOLDER:
            ret.append({
                "type": "table",
                "colLabels": [
                    "Class Level",
                    "Spell"
                ],
                "colStyles": [
                    "col-4 text-center",
                    "col-8 text-left"
                ],
                "rows": [[lvl,", ".join(['{' + f"@spell {el.lower()}" + '}' for el in s])] for lvl, s in spells.items()
                ]
        })
        elif type(el) == type('str'):
            ret.append(el)
        elif type(el) == dict:
            ret.append({
                "type": "list",
                "items": [
                    {"type": "entries", "name": k, "entries": v} for k, v in el.items()
                ]
            })
        elif type(el) == list:
            ret.append({
                "type": "list",
                "items": el
            })
    return ret

def parse_features(wb):
    ws = wb[FEATURE_SHEET]
    features = dict()
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        name = ws.cell(row=row, column=2).value
        lvl = int(ws.cell(row=row, column=1).value)
        desc = ws.cell(row=row, column=3).value
        desc_key = ws.cell(row=row, column=4).value
        desc_value = ws.cell(row=row, column=5).value
        
        if name not in features:
            features[name] = {"lvl": lvl, "content": []}

        if desc is not None:
            features[name]["content"] = features[name]["content"] + desc.strip().split("\n")

        if desc_key is not None and desc_value is not None:
            if type(features[name]["content"][-1]) == dict:
                features[name]["content"][-1][desc_key.strip()] = desc_value.strip()
            else:
                features[name]["content"].append({desc_key.strip(): desc_value.strip()})
        elif desc_value is not None:
            if type(features[name]["content"][-1]) == list:
                features[name]["content"][-1].append(desc_value.strip())
            else:
                features[name]["content"].append([desc_value.strip()])

        row += 1
    return features

def parse_subclass_spells(wb):
    spells = dict()
    ws = wb[SPELL_SHEET]
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        lvl = ws.cell(row=row, column=1).value
        spell = ws.cell(row=row, column=2).value
        manual = ws.cell(row=row, column=3).value

        if lvl not in spells:
            spells[lvl] = []

        spells[lvl].append(f"{spell.strip().lower()}" if (manual is None or manual.lower() == "phb") else f"{spell.strip().lower()}|{manual.lower()}")
        row += 1
    return spells


def parse_class_meta(wb, features, spells):
    ws = wb[DESC_SHEET]
    dnd_class = ws.cell(row=2, column=2).value
    dnd_subclass = ws.cell(row=3, column=2).value
    ret = {
        "name": dnd_subclass,
        "source": "TfI",
        "className": dnd_class,
        "classSource": "PHB",
        "shortName": ws.cell(row=4, column=2).value,
        "subclassFeatures": [
            f"{k}|{dnd_class}|PHB|{dnd_subclass}|TfI|{v['lvl']}" for k, v in features.items()
        ] 
    }
    if len(spells) > 0:
        ret["subclassSpells"] = [el for _, v in spells.items() for el in v ]
    return ret

def parse_class(wb):
    spells = parse_subclass_spells(wb)
    raw_features = parse_features(wb)
    meta = parse_class_meta(wb, raw_features, spells)
    template = { "subclass": meta, "features": [] }
    
    for f_name, f_values in raw_features.items():
        feature = {
            "name": f_name,
            "source": "TfI",
            "className": meta["className"],
            "classSource": "PHB",
            "subclassShortName": meta["name"],
            "subclassSource": "TfI",
            "level": f_values["lvl"],
            "entries": raw_features_to_feature(f_values["content"], spells)
        }
        template["features"] = template["features"] + [feature]
    return template

def parse_talent_meta(wb, features):
    ws_general = wb[DESC_SHEET]
    icon = ws_general.cell(row=5, column=2).value
    ws_talent = wb[FEATURE_SHEET]
    talent_name = ws_talent.cell(row=2, column=2).value
    feature = {
        "name": talent_name,
        "source": "TfI",
        "icon": icon,
        "entries": raw_features_to_feature(features[talent_name]["content"],[])
    }
    return feature

def parse_ability_and_prerequiste(wb):
    ws = wb[TALENTS_SHEET]
    prerequisite = list()
    ability = list()
    row = 2
    while ws.cell(row=row, column=1).value is not None or ws.cell(row=row, column=3).value is not None:
        if(ws.cell(row=row, column=1).value is not None):
            condition = ws.cell(row=row, column=1).value.strip()
            if(condition == 'other'):
                prerequisite.append({'other': ws.cell(row=row, column=2).value.strip()})
            else:
                prerequisite.append({condition: {'name': ws.cell(row=row, column=2).value.strip(), 'source': 'TfI'}})
        else:
            abi = ws.cell(row=row, column=3).value.strip()
            ability_splitted = abi.split('|')
            if(len(ability_splitted) > 1 ):
                ability.append({"choose": {"from": ability_splitted,"amount":int(ws.cell(row=row, column=4).value)}})
            else:
                ability.append({"choose": {"from": abi,"amount":int(ws.cell(row=row, column=4).value)}})
        row += 1 
    return [prerequisite,ability]


def parse_talent(wb):
    spells = parse_subclass_spells(wb)
    raw_features = parse_features(wb)
    meta = parse_talent_meta(wb, raw_features)
    [prerequisite,ability] = parse_ability_and_prerequiste(wb)
    meta["prerequisite"]=prerequisite
    meta["ability"]=ability
    return meta

language = 'en'
filenames = [f for f in os.listdir(os.path.join('source', language)) if f.endswith("xlsx")]

for file in filenames:
    wb = load_workbook(os.path.join("source",language,file))
    resource_type = wb[DESC_SHEET].cell(row=1, column=2).value.strip()
    t, resource = parse_resource_by_type(wb, resource_type)
    json.dump(resource, open(os.path.join(language,t,file.replace('xlsx', 'json')), 'w'))

